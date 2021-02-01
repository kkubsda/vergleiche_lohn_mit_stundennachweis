from openpyxl import Workbook
wb = Workbook()
current_year = 2019


def create_all_months_worksheets(year, start_month=1):
    year = str(year)
    wb.active.title = f"{month_int_to_string(start_month)}.{year}"

    for month in range(start_month + 1, 13):
        current_month = month_int_to_string(start_month + 1)
        worksheet_title = f"{current_month}.{year}"
        wb.create_sheet(worksheet_title)
        start_month += 1


def month_int_to_string(month):
    if month <= 9:
        current_month = f"0{month}"
    else:
        current_month = str(month)
    return current_month


def name_columns():
    for ws in wb:
        ws['A1'] = 'Bruttolohn'
        ws['B1'] = 'Faktor Arbeitsstunden'
        ws['C1'] = 'Arbeitsstunden ausgezahlt'
        ws['D1'] = 'Arbeitsstunden laut Stundennachweis'
        ws['E1'] = 'Lohn laut Stundennachweis'
        ws['G1'] = 'Differenz'


def insert_wages_per_hour(start_month_num, duration_in_months, hourly_wage):
    if not isinstance(start_month_num, int):
        raise ValueError('start_month_num needs to be an integer')

    for i, worksheet in enumerate(wb, 1):
        if i < start_month_num:
            continue

        worksheet['B2'] = hourly_wage

        if i == duration_in_months + start_month_num - 1:
            break


def hours_paid():
    for ws in wb:
        ws['C2'] = '=A2/B2'


def wage_according_to_record():
    for ws in wb:
        ws['E2'] = '=B2*D2'


def difference():
    for ws in wb:
        ws['G2'] = '=E2-A2'


create_all_months_worksheets(current_year)
name_columns()
hours_paid()
wage_according_to_record()
difference()

insert_wages_per_hour(1, 6, 9.19)
insert_wages_per_hour(7, 1, 12.9)
insert_wages_per_hour(8, 5, 15.9)


# change background color of the tabs
for ws in wb:
    ws.sheet_properties.tabColor = 'FF4500'

print(wb.sheetnames)
wb.save("Vergleich von erhaltenem Lohn mit Stundennachweis.xlsx")